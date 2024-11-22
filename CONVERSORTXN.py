import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# Diretórios e caminhos (substitua pelos caminhos reais no seu ambiente)
INPUT_DIR = r'CAMINHO\PARA\INPUT'
OUTPUT_DIR = r'CAMINHO\PARA\OUTPUT'
CONTROL_FILE = r'CAMINHO\PARA\ARQUIVO_DE_CONTROLE.xlsx'

# Obter a data atual
data_atual = datetime.now()

# Nome do arquivo Excel de entrada
nome_arquivo_entrada = f"TXN_{data_atual.strftime('%Y%m%d')}.xlsx"
INPUT_FILE = os.path.join(INPUT_DIR, nome_arquivo_entrada)

def get_next_batch_number():
    """
    Obtém o próximo número de lote do arquivo de controle.
    """
    try:
        wb = load_workbook(CONTROL_FILE)
        ws = wb.active
        last_row = ws.max_row
        last_batch = ws.cell(row=last_row, column=2).value
        return int(last_batch) + 1
    except Exception as e:
        print(f"Erro ao ler o número do lote: {e}")
        return 179  # Fallback para o próximo após 178

def update_control_file(batch_number, successful_records):
    """
    Atualiza o arquivo de controle com o novo lote e número de registros processados com sucesso.
    """
    try:
        wb = load_workbook(CONTROL_FILE)
        ws = wb.active
        new_row = [datetime.now().strftime('%d/%m/%Y'),
                   f'{batch_number:06d}', successful_records]
        ws.append(new_row)
        wb.save(CONTROL_FILE)
        print(f"Planilha de controle atualizada com sucesso. Novo lote: {batch_number:06d}, Registros inseridos: {successful_records}")
    except Exception as e:
        print(f"Erro ao atualizar a planilha de controle: {e}")

def create_header(batch_number: int) -> str:
    """
    Cria o cabeçalho do arquivo de saída com base no número do lote.
    """
    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    return f"000000{batch_number:03d}0000000{' ' * 40}A{current_datetime}M00000000{' ' * 412}00000002"

def create_detail_record(card_number: str, txn_code: str, value: float, date: str, sequence: int) -> str:
    """
    Cria um registro de detalhe formatado para cada linha do DataFrame.
    """
    formatted_card = f"{int(card_number):016d}"
    formatted_txn = f"{int(txn_code):04d}"
    formatted_value = f"{int(value * 100):017d}"
    formatted_date = datetime.strptime(date, '%Y-%m-%d').strftime('%Y%m%d') + "163000"

    return (
        f"1{'0' * 26}"
        f"{formatted_card}"
        f"{'0' * 7}"
        f"{formatted_txn}{formatted_txn}986"
        f"{formatted_value}"
        f"2{'0' * 17}6"
        f"{formatted_date}"
        f"{' ' * 21}"
        f"{'0' * 5}"
        f"{' ' * 79}"
        f"{'0' * 6}"
        f"{' ' * 16}"
        f"{'0' * 14}"
        f"{' ' * 23}"
        f"{'0' * 37}2"
        f"{'0' * 17}2"
        f"{'0' * 17}2"
        f"{'0' * 12}2 "
        f"{'0' * 71}"
        f"{' ' * 15}"
        f"{'0' * 8}"
        f"{' ' * 26}"
        f"{'0' * 5}2   "
        f"{sequence:08d}"
    )

def create_trailer(batch_number: int, total_records: int, total_value: int) -> str:
    """
    Cria o trailer do arquivo de saída com informações de controle.
    """
    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    return (
        f"9{'0' * 5}2{'0' * 11}"
        f"{' ' * 40}"
        f"A{current_datetime}M{current_datetime}"
        f"{total_records:08d}"
        f"{total_value:017d}"
        f"2{' ' * 378}"
        f"{'0' * 7}2"
    )

def generate_file(df: pd.DataFrame, output_path: str, batch_number: int) -> int:
    """
    Gera um arquivo de saída baseado nos dados do DataFrame, incluindo cabeçalho, registros de detalhe e trailer.
    """
    try:
        total_value = 0
        successful_records = 0
        with open(output_path, 'w') as f:
            header = create_header(batch_number)
            f.write(header + '\n')

            for i, row in df.iterrows():
                try:
                    card_number = row['NUMERO CARTÃO']
                    txn_code = row['TXN']
                    value = row['VALOR']
                    date = row['DATA DE ENVIO']

                    detail_record = create_detail_record(str(card_number), str(txn_code), value, date.strftime('%Y-%m-%d'), i + 1)
                    f.write(detail_record + '\n')
                    total_value += int(value * 100)
                    successful_records += 1
                except Exception as e:
                    print(f"Erro ao processar registro {i + 1}: {e}")

            trailer = create_trailer(batch_number, successful_records, total_value)
            f.write(trailer + '\n')

        print(f"Arquivo gerado com sucesso: {output_path}")
        return successful_records
    except Exception as e:
        raise IOError(f"Erro ao gerar o arquivo: {e}")

def main():
    """
    Função principal que coordena a leitura do arquivo de entrada, geração do arquivo de saída e atualização do controle.
    """
    try:
        if not os.path.exists(INPUT_FILE):
            raise FileNotFoundError(f"Arquivo de entrada não encontrado: {INPUT_FILE}")

        df = pd.read_excel(INPUT_FILE)

        required_columns = ['NUMERO CARTÃO', 'TXN', 'VALOR', 'DATA DE ENVIO']
        if not all(col in df.columns for col in required_columns):
            raise ValueError("Colunas necessárias não encontradas na planilha.")

        batch_number = get_next_batch_number()

        current_datetime = datetime.now()
        output_filename = f"NIO_OFFLINE_{current_datetime.strftime('%d%m%Y_%H%M')}.txt"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        successful_records = generate_file(df, output_path, batch_number)

        update_control_file(batch_number, successful_records)

        print(f"Total de registros processados com sucesso: {successful_records}")
        print(f"Total de registros na planilha original: {len(df)}")
        if successful_records < len(df):
            print(f"Atenção: {len(df) - successful_records} registros não foram processados.")

    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    main()
