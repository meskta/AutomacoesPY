import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Diretórios e credenciais genéricos
INPUT_DIR = r"C:\CAMINHO\PARA\ENTRADA"
OUTPUT_DIR = r"C:\CAMINHO\PARA\ENVIADOS"
HISTORICO_DIR = r"C:\CAMINHO\PARA\HISTORICO"
PROCESSADOS_DIR = r"C:\CAMINHO\PARA\PROCESSADOS"
CONTROL_FILE = r"C:\CAMINHO\PARA\CONTROLE_OFFLINE.xlsx"

MOVER_ARQUIVO_ORIGINAL = False

SERVIDOR_SMTP = "smtp.seuprovedor.com"
PORTA_SMTP = 465
EMAIL_PADRAO = "usuario@seudominio.com"
SENHA_PADRAO = "SUA_SENHA_AQUI"

data_atual = datetime.now()

def verificar_acesso_rede():
    diretorios = [INPUT_DIR, OUTPUT_DIR, HISTORICO_DIR, PROCESSADOS_DIR]
    for diretorio in diretorios:
        try:
            if not os.path.exists(diretorio):
                print(f"Aviso: Diretório não encontrado: {diretorio}")
                try:
                    os.makedirs(diretorio, exist_ok=True)
                    print(f"Diretório criado: {diretorio}")
                except Exception as e:
                    print(f"Erro ao criar diretório: {e}")
            else:
                os.listdir(diretorio)
                print(f"Acesso confirmado ao diretório: {diretorio}")
        except Exception as e:
            raise Exception(f"Erro ao acessar o diretório {diretorio}: {e}")

def encontrar_arquivos_txn(diretorio):
    print(f"Iniciando busca por arquivos TXN em: {diretorio}")
    try:
        todos_arquivos = os.listdir(diretorio)
        arquivos_excel = [f for f in todos_arquivos if f.lower().endswith(('.xlsx', '.xls'))]
        data_hoje = data_atual.strftime('%Y%m%d')
        padrao_hoje = f"TXN_{data_hoje}"
        arquivos_txn_hoje = [f for f in arquivos_excel if padrao_hoje.upper() in f.upper()]
        if arquivos_txn_hoje:
            print(f"Encontrados {len(arquivos_txn_hoje)} arquivos com o padrão TXN_{data_hoje}:")
            for arquivo in arquivos_txn_hoje:
                print(f"  - {arquivo}")
            return arquivos_txn_hoje
        print("Nenhum arquivo TXN do dia encontrado.")
        return []
    except Exception as e:
        print(f"Erro ao listar arquivos no diretório {diretorio}: {e}")
        return []

def get_next_batch_number():
    try:
        wb = load_workbook(CONTROL_FILE)
        ws = wb.active
        last_row = ws.max_row
        last_batch = ws.cell(row=last_row, column=2).value
        return int(last_batch) + 1
    except Exception as e:
        print(f"Erro ao ler o número do lote: {e}. Verifique se o arquivo de controle está acessível.")
        return 100

def update_control_file(batch_number, successful_records):
    try:
        wb = load_workbook(CONTROL_FILE)
        ws = wb.active
        new_row = [datetime.now().strftime('%d/%m/%Y'), f'{batch_number:06d}', successful_records]
        ws.append(new_row)
        wb.save(CONTROL_FILE)
        print(f"Planilha de controle atualizada. Novo lote: {batch_number:06d}, Registros: {successful_records}")
    except Exception as e:
        print(f"Erro ao atualizar a planilha de controle: {e}.")

def create_header(batch_number: int) -> str:
    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    return f"000000{batch_number:03d}0000000{' ' * 40}A{current_datetime}M00000000{' ' * 412}00000002"

def create_detail_record(card_number: str, txn_code: str, value: float, date: str, sequence: int) -> str:
    formatted_card = f"{int(card_number):016d}"
    formatted_txn = f"{int(txn_code):04d}"
    formatted_value = f"{int(value * 100):017d}"
    formatted_date = datetime.strptime(date, '%Y-%m-%d').strftime('%Y%m%d') + "163000"
    return (
        f"1{'0' * 26}"
        f"{formatted_card}"
        f"{'0' * 7}"
        f"{formatted_txn}{formatted_txn}986"
        f"{formatted_value}"
        f"2{'0' * 17}6"
        f"{formatted_date}"
        f"{' ' * 21}"
        f"{'0' * 5}"
        f"{' ' * 79}"
        f"{'0' * 6}"
        f"{' ' * 16}"
        f"{'0' * 14}"
        f"{' ' * 23}"
        f"{'0' * 37}2"
        f"{'0' * 17}2"
        f"{'0' * 17}2"
        f"{'0' * 12}2 "
        f"{'0' * 71}"
        f"{' ' * 15}"
        f"{'0' * 8}"
        f"{' ' * 26}"
        f"{'0' * 5}2   "
        f"{sequence:08d}"
    )

def create_trailer(batch_number: int, total_records: int, total_value: int) -> str:
    current_datetime = datetime.now().strftime('%Y%m%d%H%M%S')
    return (
        f"9{'0' * 5}2{'0' * 11}"
        f"{' ' * 40}"
        f"A{current_datetime}M{current_datetime}"
        f"{total_records:08d}"
        f"{total_value:017d}"
        f"2{' ' * 378}"
        f"{'0' * 7}2"
    )

def generate_file(df: pd.DataFrame, output_path: str, batch_number: int) -> int:
    try:
        total_value = 0
        successful_records = 0
        with open(output_path, 'w') as f:
            header = create_header(batch_number)
            f.write(header + '\n')
            for i, row in df.iterrows():
                try:
                    card_number = row['NUMERO CARTÃO']
                    txn_code = row['TXN']
                    value = row['VALOR']
                    date = row['DATA DE ENVIO'].strftime('%Y-%m-%d')
                    detail_record = create_detail_record(str(card_number), str(txn_code), value, date, i + 1)
                    f.write(detail_record + '\n')
                    total_value += int(value * 100)
                    successful_records += 1
                except Exception as e:
                    print(f"Erro ao processar registro {i + 1}: {e}")
            trailer = create_trailer(batch_number, successful_records, total_value)
            f.write(trailer + '\n')
        print(f"Arquivo gerado: {output_path}")
        return successful_records
    except Exception as e:
        raise IOError(f"Erro ao gerar o arquivo: {e}")

def enviar_email(remetente, destinatarios, assunto, corpo, senha, anexos=None):
    mensagem = MIMEMultipart()
    mensagem['From'] = remetente
    mensagem['To'] = ", ".join(destinatarios)
    mensagem['Subject'] = assunto
    mensagem.attach(MIMEText(corpo, 'html'))
    if anexos:
        for arquivo in anexos:
            if os.path.isfile(arquivo):
                with open(arquivo, 'rb') as f:
                    parte = MIMEApplication(f.read(), Name=os.path.basename(arquivo))
                parte['Content-Disposition'] = f'attachment; filename="{os.path.basename(arquivo)}"'
                mensagem.attach(parte)
    try:
        with smtplib.SMTP_SSL(SERVIDOR_SMTP, PORTA_SMTP) as servidor:
            servidor.login(remetente, senha)
            texto = mensagem.as_string()
            servidor.sendmail(remetente, destinatarios, texto)
            print("Email enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar email: {e}")
        raise

def main():
    batch_number = 0
    try:
        print(f"Iniciando processamento em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        print("Verificando acesso aos diretórios de rede...")
        verificar_acesso_rede()
        print(f"Buscando arquivos em {INPUT_DIR}...")
        arquivos_entrada = encontrar_arquivos_txn(INPUT_DIR)
        if not arquivos_entrada:
            mensagem = "Nenhum arquivo de entrada TXN encontrado."
            print(mensagem)
            remetente = EMAIL_PADRAO
            senha = SENHA_PADRAO
            destinatarios = [EMAIL_PADRAO]
            assunto = "Processamento - Arquivo OFFLINE"
            corpo = f"""
            <html>
            <body>
                <h2>Informação de Processamento</h2>
                <p>{mensagem}</p>
                <p>Data e hora da verificação: {data_atual.strftime('%d/%m/%Y %H:%M:%S')}</p>
                <p>Este é um email automático.</p>
            </body>
            </html>
            """
            try:
                enviar_email(remetente, destinatarios, assunto, corpo, senha)
            except Exception as email_error:
                print(f"Erro ao enviar email: {email_error}")
            return
        print(f"Encontrados {len(arquivos_entrada)} arquivos para processamento.")
        for arquivo in arquivos_entrada:
            INPUT_FILE = os.path.join(INPUT_DIR, arquivo)
            print(f"Processando arquivo: {INPUT_FILE}")
            if not os.path.exists(INPUT_FILE):
                raise FileNotFoundError(f"Arquivo de entrada não encontrado: {INPUT_FILE}")
            df = pd.read_excel(INPUT_FILE)
            print(f"Planilha carregada com {len(df)} registros.")
            df['DATA DE ENVIO'] = pd.to_datetime(df['DATA DE ENVIO'], errors='coerce')
            required_columns = ['NUMERO CARTÃO', 'TXN', 'VALOR', 'DATA DE ENVIO']
            if not all(col in df.columns for col in required_columns):
                raise ValueError("Colunas necessárias não encontradas na planilha.")
            batch_number = get_next_batch_number()
            print(f"Número do lote: {batch_number}")
            current_datetime = datetime.now()
            output_filename = f"OFFLINE_{current_datetime.strftime('%d%m%Y_%H%M')}.txt"
            output_path = os.path.join(OUTPUT_DIR, output_filename)
            successful_records = generate_file(df, output_path, batch_number)
            update_control_file(batch_number, successful_records)
            remetente = EMAIL_PADRAO
            senha = SENHA_PADRAO
            destinatarios = [EMAIL_PADRAO]
            assunto = f"Processamento de Transações OFFLINE_{batch_number:06d}"
            corpo = f"""
            <html>
            <body>
                <h2>Processamento Concluído</h2>
                <p>O processamento do lote OFFLINE <b>{batch_number:06d}</b> foi concluído.</p>
                <ul>
                    <li>Arquivo processado: {arquivo}</li>
                    <li>Arquivo gerado: {output_filename}</li>
                    <li>Total de registros processados: {successful_records}</li>
                    <li>Total de registros na planilha original: {len(df)}</li>
                    <li>Data e hora do processamento: {current_datetime.strftime('%d/%m/%Y %H:%M:%S')}</li>
                </ul>
                <p>Este é um email automático.</p>
            </body>
            </html>
            """
            anexos = [output_path]
            try:
                enviar_email(remetente, destinatarios, assunto, corpo, senha, anexos)
            except Exception as email_error:
                print(f"Erro ao enviar email: {email_error}")
            print(f"Total de registros processados com sucesso: {successful_records}")
            print(f"Total de registros na planilha original: {len(df)}")
            if successful_records < len(df):
                print(f"Atenção: {len(df) - successful_records} registros não foram processados.")
            data_hoje = current_datetime.strftime('%Y%m%d')
            arquivo_txn_historico = f"TXN_{data_hoje}.txt"
            destino_txn = os.path.join(HISTORICO_DIR, arquivo_txn_historico)
            if MOVER_ARQUIVO_ORIGINAL:
                shutil.move(output_path, destino_txn)
                print(f"Arquivo TXN movido para o histórico: {destino_txn}")
            else:
                shutil.copy2(output_path, destino_txn)
                print(f"Arquivo TXN mantido em: {output_path}")
                print(f"Cópia do arquivo TXN salva em: {destino_txn}")
            if not os.path.exists(PROCESSADOS_DIR):
                os.makedirs(PROCESSADOS_DIR, exist_ok=True)
            destino_excel = os.path.join(PROCESSADOS_DIR, arquivo)
            shutil.move(INPUT_FILE, destino_excel)
            print(f"Arquivo Excel movido para: {destino_excel}")
        print(f"Processamento concluído em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    except Exception as e:
        erro_msg = f"Erro: {e}"
        print(erro_msg)
        try:
            remetente = EMAIL_PADRAO
            senha = SENHA_PADRAO
            destinatarios = [EMAIL_PADRAO]
            assunto = "ERRO - Processamento de Transações OFFLINE"
            corpo = f"""
            <html>
            <body>
                <h2>Log de processamento - OFFLINE_{batch_number:06d}</h2>
                <p>Resultado do processamento:</p>
                <p style="color: red; font-weight: bold;">{str(e)}</p>
                <p>Data e hora do erro: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
                <ul>
                    <li>Verifique se o arquivo TXN está presente no diretório de entrada</li>
                    <li>Verifique se as unidades de rede estão mapeadas</li>
                    <li>Verifique os logs para mais detalhes sobre o erro</li>
                </ul>
                <p>Este é um email automático.</p>
            </body>
            </html>
            """
            try:
                enviar_email(remetente, destinatarios, assunto, corpo, senha)
            except Exception as email_error:
                print(f"Erro ao enviar email de notificação de erro: {email_error}")
        except Exception as email_error:
            print(f"Erro ao enviar email de notificação de erro:
